from openpyxl import load_workbook
from openpyxl import Workbook

class Comparison:
    sophos_workbook = None
    continuum_workbook = None
    sophos_sheet = None
    continuum_sheet = None
    sophos_max_rows = None
    continuum_max_rows = None
    sophos_not_on_continuum_workbook = Workbook()
    sophos_not_on_continuum_sheet = sophos_not_on_continuum_workbook.active
    sophos_not_on_continuum_sheet.cell(row=1,column=1).value = "Endpoint"
    sophos_not_on_continuum_sheet.cell(row=1,column=2).value = "Site"


    def load_workbooks(self, sophos_workbook, continuum_workbook):
        self.sophos_workbook = sophos_workbook
        self.continuum_workbook = continuum_workbook
        self.sophos_sheet = sophos_workbook.active
        self.continuum_sheet = continuum_workbook.active

    def get_maximum_rows(self):
        rows = 0
        for max_row, row in enumerate(self.sophos_sheet, 1):
            if not all(col.value is None for col in row):
                rows += 1
        self.sophos_max_rows = rows

        rows = 0

        for max_row, row in enumerate(self.continuum_sheet, 1):
            if not all(col.value is None for col in row):
                rows += 1
        self.continuum_max_rows = rows

    def compare_sophos_to_continuum(self):
        self.get_maximum_rows()
        count = 2
        for val in range(self.sophos_max_rows):
            sophos_cell = self.sophos_sheet.cell(row=val+1,column=2)
            for valu in range(self.continuum_max_rows):
                continuum_cell = self.continuum_sheet.cell(row=valu+1,column=1)
                if continuum_cell.value == sophos_cell.value:
                    break
                if valu + 1 == self.continuum_max_rows:
                    self.sophos_not_on_continuum_sheet.cell(row=count,column=1).value = self.sophos_sheet.cell(row=val+1,column=2).value
                    self.sophos_not_on_continuum_sheet.cell(row=count,column=2).value = self.sophos_sheet.cell(row=val+1,column=1).value
                    count += 1
                    self.sophos_not_on_continuum_workbook.save("test.xlsx")
            
                
            
            
            


